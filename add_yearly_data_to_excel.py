import pandas as pd
import os
import glob
from openpyxl import load_workbook

# 出力ディレクトリ
data_dir = "C:\\Users\\rilak\\Desktop\\株価\\株価データ"
print(f"処理対象ディレクトリ: {data_dir}")

# 年足CSVファイルを検索
yearly_csv_files = glob.glob(os.path.join(data_dir, "*_年足.csv"))
print(f"見つかった年足CSVファイル: {len(yearly_csv_files)}個")

for yearly_csv_file in yearly_csv_files:
    try:
        # ファイル名から情報を抽出
        basename = os.path.basename(yearly_csv_file)
        ticker_and_name = basename.replace('_年足.csv', '')
        
        print(f"処理中: {basename} -> {ticker_and_name}")
        
        # 元のExcelファイルのパス
        excel_file = os.path.join(data_dir, f"{ticker_and_name}.xlsx")
        
        if not os.path.exists(excel_file):
            print(f"警告: 元のExcelファイルが見つかりません: {excel_file}")
            continue
        
        # 年足CSVデータを読み込む
        yearly_data = pd.read_csv(yearly_csv_file, index_col=0, parse_dates=True)
        
        # ティッカーシンボルまたは銘柄コードから銘柄名を抽出
        # ファイル名のパターンは通常「コード_銘柄名_年足.csv」の形式と想定
        # または単に「銘柄名_年足.csv」の場合もある
        
        # ファイル名から銘柄名を抽出
        if "_" in ticker_and_name:
            # ファイル名に_が含まれる場合、最後の部分を銘柄名と仮定
            stock_name = ticker_and_name.split("_")[-1]
        else:
            # _がない場合はそのままを銘柄名として使用
            stock_name = ticker_and_name
            
        # シート名を作成（銘柄名_年足）
        sheet_name = f"{stock_name}_年足"
        
        # シート名の長さ制限（31文字まで）
        sheet_name = sheet_name[:31]
        
        # シートの存在を確認し、存在する場合は上書き
        try:
            book = load_workbook(excel_file)
            if sheet_name in book.sheetnames:
                print(f"シート '{sheet_name}' はすでに存在します。削除します。")
                std = book[sheet_name]
                book.remove(std)
                book.save(excel_file)
            book.close()
        except Exception as e:
            print(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
            continue
        
        try:
            # Excelファイルに年足データを追加
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                yearly_data.to_excel(writer, sheet_name=sheet_name)
            print(f"年足データを元のExcelファイルに追加しました: {excel_file} (シート: {sheet_name})")
        except Exception as e:
            print(f"エラー: {e}")
    
    except Exception as e:
        print(f"処理中にエラーが発生しました: {e}")
    
    print("=" * 50)

print("\n処理が完了しました。年足データを元のExcelファイルに追加しました。")
