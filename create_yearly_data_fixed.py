import pandas as pd
import os
import glob
from datetime import datetime
from openpyxl import load_workbook
import time

# 出力ディレクトリを確保
output_dir = "C:\\Users\\rilak\\Desktop\\株価\\株価データ"
os.makedirs(output_dir, exist_ok=True)
print(f"データをローカルフォルダに保存します: {output_dir}")

# 月足CSVファイルを検索（*_月足.csvというパターンを検索）
monthly_files = glob.glob(os.path.join(output_dir, "*_月足.csv"))
print(f"見つかった月足ファイル: {len(monthly_files)}個")

# ユーザーにカラム名の言語を確認
use_japanese_columns = input("カラムの表示を日本語にしますか？(y/n): ").strip().lower() == 'y'

for monthly_file in monthly_files:
    print(f"処理中: {os.path.basename(monthly_file)}")
    
    try:
        # ファイル名から情報を抽出（銘柄コードと銘柄名）
        basename = os.path.basename(monthly_file)
        base_name_without_ext = os.path.splitext(basename)[0]  # 拡張子なしのファイル名
        ticker_and_name = base_name_without_ext.replace('_月足', '')  # 月足部分を削除
        
        # 月足データを読み込む
        df_monthly = pd.read_csv(monthly_file, index_col=0, parse_dates=True)
        
        # データがあるか確認
        if df_monthly.empty:
            print(f"データがありません: {basename}")
            continue
        
        # データ列があるか確認
        if len(df_monthly.columns) == 0:
            print(f"警告: {basename} には列がありません。この問題を修正します。")
            # ファイルを直接開いて内容を確認
            with open(monthly_file, 'r', encoding='utf-8') as f:
                content = f.read()
                print(f"ファイル内容のプレビュー: {content[:200]}...")
            
            # 再度読み込み試行（パースオプションを変更）
            df_monthly = pd.read_csv(monthly_file, parse_dates=True)
            
            # インデックスを設定
            if '日付' in df_monthly.columns:
                df_monthly.set_index('日付', inplace=True)
            elif 'Date' in df_monthly.columns:
                df_monthly.set_index('Date', inplace=True)
            else:
                # インデックス列を自動検出
                if df_monthly.columns[0].lower() in ['date', '日付', 'datetime', 'time', '時間']:
                    df_monthly.set_index(df_monthly.columns[0], inplace=True)
        
        # インデックスが日付型でない場合は変換
        if not isinstance(df_monthly.index, pd.DatetimeIndex):
            try:
                df_monthly.index = pd.to_datetime(df_monthly.index)
            except ValueError as e:
                if "Tz-aware datetime" in str(e):
                    # タイムゾーン情報付きの場合はUTCとして処理
                    print("タイムゾーン情報付きの日付を処理中...")
                    df_monthly.index = pd.to_datetime(df_monthly.index, utc=True)
                else:
                    raise
        
        # タイムゾーン情報を削除（Excelの互換性のため）
        if hasattr(df_monthly.index, 'tz') and df_monthly.index.tz is not None:
            df_monthly.index = df_monthly.index.tz_localize(None)
        
        # 年足データに変換
        print("年足データに変換中...")
        
        # 年ごとにグループ化（YEを使用）
        yearly_groups = df_monthly.groupby(pd.Grouper(freq='YE'))
        
        # 年足データフレームを作成
        yearly_data = pd.DataFrame()
        
        # 日本語と英語の列名を定義
        jp_columns = ['始値', '高値', '安値', '終値', '出来高', '配当', '株式分割']
        en_columns = ['Open', 'High', 'Low', 'Close', 'Volume', 'Dividends', 'Stock Splits']
        
        # 現在の列名を確認
        current_columns = list(df_monthly.columns)
        print(f"現在の列名: {current_columns}")
        
        # 日本語列名か英語列名かを判断
        is_jp = any(col in jp_columns for col in current_columns)
        
        # 列名のマッピングを作成
        if is_jp:
            col_open = '始値' if '始値' in current_columns else 'Open'
            col_high = '高値' if '高値' in current_columns else 'High'
            col_low = '安値' if '安値' in current_columns else 'Low'
            col_close = '終値' if '終値' in current_columns else 'Close'
            col_volume = '出来高' if '出来高' in current_columns else 'Volume'
            col_dividends = '配当' if '配当' in current_columns else 'Dividends'
            col_splits = '株式分割' if '株式分割' in current_columns else 'Stock Splits'
        else:
            col_open = 'Open' if 'Open' in current_columns else '始値'
            col_high = 'High' if 'High' in current_columns else '高値'
            col_low = 'Low' if 'Low' in current_columns else '安値'
            col_close = 'Close' if 'Close' in current_columns else '終値'
            col_volume = 'Volume' if 'Volume' in current_columns else '出来高'
            col_dividends = 'Dividends' if 'Dividends' in current_columns else '配当'
            col_splits = 'Stock Splits' if 'Stock Splits' in current_columns else '株式分割'
        
        # グループごとに年足データを計算
        yearly_data_list = []
        for name, group in yearly_groups:
            if not group.empty:
                # 年足データの計算
                yearly_row = {}
                
                # 各カラムの処理を安全に行う
                if col_open in group.columns:
                    yearly_row[col_open] = group.iloc[0][col_open]  # 年初の始値
                
                if col_high in group.columns:
                    yearly_row[col_high] = group[col_high].max()    # 年間最高値
                
                if col_low in group.columns:
                    yearly_row[col_low] = group[col_low].min()      # 年間最安値
                
                if col_close in group.columns:
                    yearly_row[col_close] = group.iloc[-1][col_close]  # 年末の終値
                
                if col_volume in group.columns:
                    yearly_row[col_volume] = group[col_volume].sum()  # 年間出来高合計
                
                if col_dividends in group.columns:
                    yearly_row[col_dividends] = group[col_dividends].sum()  # 年間配当合計
                
                if col_splits in group.columns:
                    yearly_row[col_splits] = group[col_splits].sum()  # 年間株式分割合計
                
                yearly_data_list.append((name, yearly_row))
        
        # DataFrameに変換
        if yearly_data_list:
            yearly_data = pd.DataFrame([row for _, row in yearly_data_list], index=[name for name, _ in yearly_data_list])
        
        # カラム名を希望の言語に変更
        if use_japanese_columns and not is_jp:
            column_mapping = {
                'Open': '始値',
                'High': '高値',
                'Low': '安値',
                'Close': '終値',
                'Volume': '出来高',
                'Dividends': '配当',
                'Stock Splits': '株式分割'
            }
            # 存在する列名のみをマッピング
            mapping = {col: column_mapping[col] for col in yearly_data.columns if col in column_mapping}
            yearly_data.rename(columns=mapping, inplace=True)
        elif not use_japanese_columns and is_jp:
            column_mapping = {
                '始値': 'Open',
                '高値': 'High',
                '安値': 'Low',
                '終値': 'Close',
                '出来高': 'Volume',
                '配当': 'Dividends',
                '株式分割': 'Stock Splits'
            }
            # 存在する列名のみをマッピング
            mapping = {col: column_mapping[col] for col in yearly_data.columns if col in column_mapping}
            yearly_data.rename(columns=mapping, inplace=True)
        
        # インデックス名を設定
        yearly_data.index.name = '日付' if use_japanese_columns else 'Date'
        
        # 年足CSVファイルを保存
        yearly_csv_path = os.path.join(output_dir, f"{ticker_and_name}_年足.csv")
        yearly_data.to_csv(yearly_csv_path)
        print(f"年足CSVファイルを保存しました: {yearly_csv_path}")
        
        # 元のExcelファイルに年足データを追加
        combined_excel_path = os.path.join(output_dir, f"{ticker_and_name}.xlsx")
        
        # シート名の決定（銘柄名_年足）
        if "任天堂" in ticker_and_name:
            sheet_name = "任天堂_年足"
        elif "eMAXIS" in ticker_and_name:
            sheet_name = "eMAXIS_年足"
        elif "アップル" in ticker_and_name:
            sheet_name = "アップル_年足"
        else:
            sheet_name = f"{ticker_and_name.split('_')[-1][:15]}_年足"
        
        # シート名の長さ制限（31文字まで）
        sheet_name = sheet_name[:31]
        
        if os.path.exists(combined_excel_path):
            # ファイルが開かれていないことを確認
            max_attempts = 5
            for attempt in range(max_attempts):
                try:
                    # シートの存在を確認し、存在する場合は削除
                    book = load_workbook(combined_excel_path)
                    if sheet_name in book.sheetnames:
                        print(f"シート '{sheet_name}' はすでに存在します。削除します。")
                        std = book[sheet_name]
                        book.remove(std)
                        book.save(combined_excel_path)
                    book.close()
                    
                    # ファイルの保存と閉じる操作の後に少し待機
                    time.sleep(1)
                    
                    # Excelファイルに年足データを追加
                    with pd.ExcelWriter(combined_excel_path, engine='openpyxl', mode='a') as writer:
                        yearly_data.to_excel(writer, sheet_name=sheet_name)
                    print(f"年足データを元のExcelファイルに追加しました: {combined_excel_path} (シート: {sheet_name})")
                    break  # 成功したらループを抜ける
                    
                except PermissionError as e:
                    if attempt < max_attempts - 1:
                        print(f"ファイルが使用中です。再試行します... ({attempt+1}/{max_attempts})")
                        time.sleep(2)  # 2秒待機
                    else:
                        print(f"権限エラー (最大試行回数に到達): {e}")
                        print("年足データを元のExcelファイルに追加できませんでした。")
                except Exception as e:
                    print(f"元のExcelファイルに年足データを追加できませんでした: {e}")
                    break
        else:
            # Excelファイルが存在しない場合は新規作成
            with pd.ExcelWriter(combined_excel_path, engine='openpyxl') as writer:
                yearly_data.to_excel(writer, sheet_name=sheet_name)
            print(f"新しいExcelファイルを作成し、年足データを追加しました: {combined_excel_path} (シート: {sheet_name})")
        
        # データの最初と最後の行を表示
        if not yearly_data.empty:
            print("年足データの最初の行:")
            print(yearly_data.head(1))
            
            print("年足データの最後の行:")
            print(yearly_data.tail(1))
        else:
            print("警告: 年足データが空です。")
        
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        import traceback
        print(traceback.format_exc())  # 詳細なエラー情報を表示
    
    print("\n" + "="*80 + "\n")  # 区切り線

print("処理が完了しました。年足データを作成し、保存しました。")
