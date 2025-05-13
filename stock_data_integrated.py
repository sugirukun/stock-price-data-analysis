import yfinance as yf
import pandas as pd
import os
import time
from datetime import datetime
import glob
from openpyxl import load_workbook

def main():
    print("===== 株価データ総合処理ツール =====")
    print("このスクリプトは以下の処理を統合しています：")
    print("1. 株価データの取得 (stock_data_all.py)")
    print("2. 年足データの作成 (create_yearly_data_fixed.py)")
    print("3. 年足データのExcelファイルへの追加 (add_yearly_data_to_excel.py)")
    print("=" * 50)

    # 処理の選択
    print("\n実行する処理を選択してください：")
    print("1. すべての処理を順番に実行")
    print("2. 株価データの取得のみ")
    print("3. 年足データの作成のみ")
    print("4. 年足データのExcelファイルへの追加のみ")
    
    choice = input("選択してください (1-4): ").strip()
    
    # 出力ディレクトリを確保
    output_dir = "C:\\Users\\rilak\\Desktop\\株価\\株価データ"
    os.makedirs(output_dir, exist_ok=True)
    print(f"\nデータをローカルフォルダに保存します: {output_dir}")
    
    # カラム名の日本語表示の選択
    use_japanese_columns = input("\nカラムの表示を日本語にしますか？(y/n)\n※日本語表記にすることでこのプログラムの実行は問題は出ませんが、より詳細な分析をするときは支障が出ることがあります: ").strip().lower() == 'y'
    
    # 選択に基づいて処理を実行
    if choice == '1':
        # すべての処理を順番に実行
        fetch_stock_data(output_dir, use_japanese_columns)
        create_yearly_data(output_dir, use_japanese_columns)
        add_yearly_data_to_excel(output_dir)
    elif choice == '2':
        # 株価データの取得のみ
        fetch_stock_data(output_dir, use_japanese_columns)
    elif choice == '3':
        # 年足データの作成のみ
        create_yearly_data(output_dir, use_japanese_columns)
    elif choice == '4':
        # 年足データのExcelファイルへの追加のみ
        add_yearly_data_to_excel(output_dir)
    else:
        print("無効な選択です。プログラムを終了します。")
        return
    
    print("\n全ての処理が完了しました。")

def fetch_stock_data(output_dir, use_japanese_columns):
    """株価データを取得する関数（stock_data_all.pyと同等）"""
    print("\n===== 株価データの取得を開始 =====")
    
    # 取得したい銘柄のリスト
    tickers = {
        "7974.T": "任天堂",         # 任天堂の証券コード (東証)
        "1387.T": "eMAXIS Slim米国株式(S&P500)",  # eMAXIS Slim米国株式(S&P500)のETFコード
        "AAPL": "アップル"          # Apple (米国株)
    }
    
    # 期間タイプの定義
    frequency_types = {
        "日足": "1d",
        "週足": "1wk",
        "月足": "1mo",
        "年足": "1y"  # yfinanceは年足も直接取得可能
    }
    
    # 各銘柄について株価データを取得
    for ticker, name in tickers.items():
        print(f"\n{ticker}（{name}）の株価データを取得中...")
        
        try:
            # yfinanceで銘柄オブジェクトを取得
            stock = yf.Ticker(ticker)
            
            # 各期間タイプのデータを取得
            period_data = {}
            for period_name, period_code in frequency_types.items():
                print(f"{period_name}データを取得中...")
                data = stock.history(period="max", interval=period_code)
                
                # データが空でないか確認
                if data.empty:
                    print(f"{period_name}のデータがありません: {ticker}")
                    continue
                
                # データの行数と期間を表示
                start_date = data.index[0].strftime('%Y-%m-%d')
                end_date = data.index[-1].strftime('%Y-%m-%d')
                print(f"{period_name}の取得期間: {start_date}から{end_date}まで")
                print(f"{period_name}のデータ点数: {len(data)}日分")
                
                # 期間データを保存
                period_data[period_name] = data
            
            # 安全なファイル名を生成（ティッカー記号からピリオドを除去）
            safe_ticker = ticker.replace('.', '_')
            
            # カラム名を日本語に変更
            japanese_columns = {
                'Open': '始値',
                'High': '高値',
                'Low': '安値',
                'Close': '終値',
                'Volume': '出来高',
                'Dividends': '配当',
                'Stock Splits': '株式分割'
            }
            
            # CSVファイルとして各期間データを保存
            for period_name, data in period_data.items():
                # データのコピーを作成
                processed_data = data.copy()
                
                # ユーザー選択に基づいてカラム名を変更
                if use_japanese_columns:
                    processed_data.rename(columns=japanese_columns, inplace=True)
                    print(f"{period_name}のカラム名を日本語に変換しました。")
                else:
                    print(f"{period_name}のカラム名は英語のままです。")
                
                # CSVファイルとして保存
                csv_path = os.path.join(output_dir, f"{safe_ticker}_{name}_{period_name}.csv")
                processed_data.to_csv(csv_path)
                print(f"{period_name}のCSVファイルを保存しました: {csv_path}")
                
                # データの最初と最後の5行を表示
                column_type = "日本語カラム" if use_japanese_columns else "英語カラム"
                print(f"{period_name}の最初の5日分のデータ（{column_type}）:")
                print(processed_data.head())
                
                print(f"{period_name}の最新の5日分のデータ（{column_type}）:")
                print(processed_data.tail())
                print("\n" + "-"*50 + "\n")  # 区切り線
            
            # Excelファイルとして全期間データを1つのファイルに保存
            excel_path = os.path.join(output_dir, f"{safe_ticker}_{name}.xlsx")
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for period_name, data in period_data.items():
                    # データのコピーを作成
                    excel_data = data.copy()
                    
                    # ユーザー選択に基づいてカラム名を変更
                    if use_japanese_columns:
                        excel_data.rename(columns=japanese_columns, inplace=True)
                    
                    # タイムゾーン情報を削除
                    excel_data.index = excel_data.index.tz_localize(None)
                    
                    # インデックスの名前を変更
                    excel_data.index.name = '日付' if use_japanese_columns else 'Date'
                    
                    # シート名（31文字以内に制限）
                    sheet_name = f"{name[:15]}_{period_name}" if len(name) > 15 else f"{name}_{period_name}"
                    sheet_name = sheet_name[:31]  # シート名の制限のため31文字まで
                    
                    # シートにデータを書き込み
                    excel_data.to_excel(writer, sheet_name=sheet_name)
                    
                print(f"Excelファイルを保存しました（全期間データ）: {excel_path}")
            
        except Exception as e:
            print(f"エラーが発生しました: {e}")
        
        # 連続リクエストによるAPIの制限を避けるため少し待つ
        time.sleep(1)
        
        print("\n" + "="*80 + "\n")  # 区切り線
    
    print("株価データの取得が完了しました。")

def create_yearly_data(output_dir, use_japanese_columns):
    """月足データから年足データを作成する関数（create_yearly_data_fixed.pyと同等）"""
    print("\n===== 年足データの作成を開始 =====")
    
    # 月足CSVファイルを検索（*_月足.csvというパターンを検索）
    monthly_files = glob.glob(os.path.join(output_dir, "*_月足.csv"))
    print(f"見つかった月足ファイル: {len(monthly_files)}個")
    
    for monthly_file in monthly_files:
        print(f"\n処理中: {os.path.basename(monthly_file)}")
        
        try:
            # ファイル名から情報を抽出（銘柄コードと銘柄名）
            basename = os.path.basename(monthly_file)
            base_name_without_ext = os.path.splitext(basename)[0]  # 拡張子なしのファイル名
            ticker_and_name = base_name_without_ext.replace('_月足', '')  # 月足部分を削除
            
            # 月足データを読み込む
            df_monthly = pd.read_csv(monthly_file, index_col=0, parse_dates=True)
            
            # データがあるか確認
            if df_monthly.empty:
                print(f"データがありません: {basename}")
                continue
            
            # データ列があるか確認
            if len(df_monthly.columns) == 0:
                print(f"警告: {basename} には列がありません。この問題を修正します。")
                # ファイルを直接開いて内容を確認
                with open(monthly_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    print(f"ファイル内容のプレビュー: {content[:200]}...")
                
                # 再度読み込み試行（パースオプションを変更）
                df_monthly = pd.read_csv(monthly_file, parse_dates=True)
                
                # インデックスを設定
                if '日付' in df_monthly.columns:
                    df_monthly.set_index('日付', inplace=True)
                elif 'Date' in df_monthly.columns:
                    df_monthly.set_index('Date', inplace=True)
                else:
                    # インデックス列を自動検出
                    if df_monthly.columns[0].lower() in ['date', '日付', 'datetime', 'time', '時間']:
                        df_monthly.set_index(df_monthly.columns[0], inplace=True)
            
            # インデックスが日付型でない場合は変換
            if not isinstance(df_monthly.index, pd.DatetimeIndex):
                try:
                    df_monthly.index = pd.to_datetime(df_monthly.index)
                except ValueError as e:
                    if "Tz-aware datetime" in str(e):
                        # タイムゾーン情報付きの場合はUTCとして処理
                        print("タイムゾーン情報付きの日付を処理中...")
                        df_monthly.index = pd.to_datetime(df_monthly.index, utc=True)
                    else:
                        raise
            
            # タイムゾーン情報を削除（Excelの互換性のため）
            if hasattr(df_monthly.index, 'tz') and df_monthly.index.tz is not None:
                df_monthly.index = df_monthly.index.tz_localize(None)
            
            # 年足データに変換
            print("年足データに変換中...")
            
            # 年ごとにグループ化（YEを使用）
            yearly_groups = df_monthly.groupby(pd.Grouper(freq='YE'))
            
            # 年足データフレームを作成
            yearly_data = pd.DataFrame()
            
            # 日本語と英語の列名を定義
            jp_columns = ['始値', '高値', '安値', '終値', '出来高', '配当', '株式分割']
            en_columns = ['Open', 'High', 'Low', 'Close', 'Volume', 'Dividends', 'Stock Splits']
            
            # 現在の列名を確認
            current_columns = list(df_monthly.columns)
            print(f"現在の列名: {current_columns}")
            
            # 日本語列名か英語列名かを判断
            is_jp = any(col in jp_columns for col in current_columns)
            
            # 列名のマッピングを作成
            if is_jp:
                col_open = '始値' if '始値' in current_columns else 'Open'
                col_high = '高値' if '高値' in current_columns else 'High'
                col_low = '安値' if '安値' in current_columns else 'Low'
                col_close = '終値' if '終値' in current_columns else 'Close'
                col_volume = '出来高' if '出来高' in current_columns else 'Volume'
                col_dividends = '配当' if '配当' in current_columns else 'Dividends'
                col_splits = '株式分割' if '株式分割' in current_columns else 'Stock Splits'
            else:
                col_open = 'Open' if 'Open' in current_columns else '始値'
                col_high = 'High' if 'High' in current_columns else '高値'
                col_low = 'Low' if 'Low' in current_columns else '安値'
                col_close = 'Close' if 'Close' in current_columns else '終値'
                col_volume = 'Volume' if 'Volume' in current_columns else '出来高'
                col_dividends = 'Dividends' if 'Dividends' in current_columns else '配当'
                col_splits = 'Stock Splits' if 'Stock Splits' in current_columns else '株式分割'
            
            # グループごとに年足データを計算
            yearly_data_list = []
            for name, group in yearly_groups:
                if not group.empty:
                    # 年足データの計算
                    yearly_row = {}
                    
                    # 各カラムの処理を安全に行う
                    if col_open in group.columns:
                        yearly_row[col_open] = group.iloc[0][col_open]  # 年初の始値
                    
                    if col_high in group.columns:
                        yearly_row[col_high] = group[col_high].max()    # 年間最高値
                    
                    if col_low in group.columns:
                        yearly_row[col_low] = group[col_low].min()      # 年間最安値
                    
                    if col_close in group.columns:
                        yearly_row[col_close] = group.iloc[-1][col_close]  # 年末の終値
                    
                    if col_volume in group.columns:
                        yearly_row[col_volume] = group[col_volume].sum()  # 年間出来高合計
                    
                    if col_dividends in group.columns:
                        yearly_row[col_dividends] = group[col_dividends].sum()  # 年間配当合計
                    
                    if col_splits in group.columns:
                        yearly_row[col_splits] = group[col_splits].sum()  # 年間株式分割合計
                    
                    yearly_data_list.append((name, yearly_row))
            
            # DataFrameに変換
            if yearly_data_list:
                yearly_data = pd.DataFrame([row for _, row in yearly_data_list], index=[name for name, _ in yearly_data_list])
            
            # カラム名を希望の言語に変更
            if use_japanese_columns and not is_jp:
                column_mapping = {
                    'Open': '始値',
                    'High': '高値',
                    'Low': '安値',
                    'Close': '終値',
                    'Volume': '出来高',
                    'Dividends': '配当',
                    'Stock Splits': '株式分割'
                }
                # 存在する列名のみをマッピング
                mapping = {col: column_mapping[col] for col in yearly_data.columns if col in column_mapping}
                yearly_data.rename(columns=mapping, inplace=True)
            elif not use_japanese_columns and is_jp:
                column_mapping = {
                    '始値': 'Open',
                    '高値': 'High',
                    '安値': 'Low',
                    '終値': 'Close',
                    '出来高': 'Volume',
                    '配当': 'Dividends',
                    '株式分割': 'Stock Splits'
                }
                # 存在する列名のみをマッピング
                mapping = {col: column_mapping[col] for col in yearly_data.columns if col in column_mapping}
                yearly_data.rename(columns=mapping, inplace=True)
            
            # インデックス名を設定
            yearly_data.index.name = '日付' if use_japanese_columns else 'Date'
            
            # 年足CSVファイルを保存
            yearly_csv_path = os.path.join(output_dir, f"{ticker_and_name}_年足.csv")
            yearly_data.to_csv(yearly_csv_path)
            print(f"年足CSVファイルを保存しました: {yearly_csv_path}")
            
            # 元のExcelファイルに年足データを追加
            combined_excel_path = os.path.join(output_dir, f"{ticker_and_name}.xlsx")
            
            # シート名の決定（銘柄名_年足）
            if "任天堂" in ticker_and_name:
                sheet_name = "任天堂_年足"
            elif "eMAXIS" in ticker_and_name:
                sheet_name = "eMAXIS_年足"
            elif "アップル" in ticker_and_name:
                sheet_name = "アップル_年足"
            else:
                sheet_name = f"{ticker_and_name.split('_')[-1][:15]}_年足"
            
            # シート名の長さ制限（31文字まで）
            sheet_name = sheet_name[:31]
            
            if os.path.exists(combined_excel_path):
                try:
                    # シートの存在を確認し、存在する場合は削除
                    book = load_workbook(combined_excel_path)
                    if sheet_name in book.sheetnames:
                        print(f"シート '{sheet_name}' はすでに存在します。削除します。")
                        std = book[sheet_name]
                        book.remove(std)
                        book.save(combined_excel_path)
                    book.close()
                    
                    # ファイルの保存と閉じる操作の後に少し待機
                    time.sleep(1)
                    
                    # Excelファイルに年足データを追加
                    with pd.ExcelWriter(combined_excel_path, engine='openpyxl', mode='a') as writer:
                        yearly_data.to_excel(writer, sheet_name=sheet_name)
                    print(f"年足データを元のExcelファイルに追加しました: {combined_excel_path} (シート: {sheet_name})")
                    
                except PermissionError as e:
                    print(f"権限エラー: {e}")
                    print("年足データを元のExcelファイルに追加できませんでした。後で手動で追加してください。")
                except Exception as e:
                    print(f"元のExcelファイルに年足データを追加できませんでした: {e}")
            else:
                # Excelファイルが存在しない場合は新規作成
                with pd.ExcelWriter(combined_excel_path, engine='openpyxl') as writer:
                    yearly_data.to_excel(writer, sheet_name=sheet_name)
                print(f"新しいExcelファイルを作成し、年足データを追加しました: {combined_excel_path} (シート: {sheet_name})")
            
            # データの最初と最後の行を表示
            if not yearly_data.empty:
                print("年足データの最初の行:")
                print(yearly_data.head(1))
                
                print("年足データの最後の行:")
                print(yearly_data.tail(1))
            else:
                print("警告: 年足データが空です。")
            
        except Exception as e:
            print(f"エラーが発生しました: {e}")
            import traceback
            print(traceback.format_exc())  # 詳細なエラー情報を表示
        
        print("\n" + "="*80 + "\n")  # 区切り線
    
    print("年足データの作成が完了しました。")

def add_yearly_data_to_excel(output_dir):
    """年足データをExcelファイルに追加する関数（add_yearly_data_to_excel.pyと同等）"""
    print("\n===== 年足データのExcelファイルへの追加を開始 =====")
    
    # 年足CSVファイルを検索
    yearly_csv_files = glob.glob(os.path.join(output_dir, "*_年足.csv"))
    print(f"見つかった年足CSVファイル: {len(yearly_csv_files)}個")
    
    for yearly_csv_file in yearly_csv_files:
        try:
            # ファイル名から情報を抽出
            basename = os.path.basename(yearly_csv_file)
            ticker_and_name = basename.replace('_年足.csv', '')
            
            print(f"\n処理中: {basename} -> {ticker_and_name}")
            
            # 元のExcelファイルのパス
            excel_file = os.path.join(output_dir, f"{ticker_and_name}.xlsx")
            
            if not os.path.exists(excel_file):
                print(f"警告: 元のExcelファイルが見つかりません: {excel_file}")
                continue
            
            # 年足CSVデータを読み込む
            yearly_data = pd.read_csv(yearly_csv_file, index_col=0, parse_dates=True)
            
            # シート名の生成（銘柄名_年足）
            if "任天堂" in ticker_and_name:
                sheet_name = "任天堂_年足"
            elif "eMAXIS" in ticker_and_name:
                sheet_name = "eMAXIS_年足"
            elif "アップル" in ticker_and_name:
                sheet_name = "アップル_年足"
            else:
                sheet_name = f"{ticker_and_name.split('_')[-1][:15]}_年足"
            
            # シート名の長さ制限（31文字まで）
            sheet_name = sheet_name[:31]
            
            # シートの存在を確認し、存在する場合は上書き
            try:
                book = load_workbook(excel_file)
                if sheet_name in book.sheetnames:
                    print(f"シート '{sheet_name}' はすでに存在します。削除します。")
                    std = book[sheet_name]
                    book.remove(std)
                    book.save(excel_file)
                book.close()
                
                # ファイルの保存と閉じる操作の後に少し待機
                time.sleep(1)
            except Exception as e:
                print(f"Excelファイルの読み込み中にエラーが発生しました: {e}")
                continue
            
            try:
                # Excelファイルに年足データを追加
                with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                    yearly_data.to_excel(writer, sheet_name=sheet_name)
                print(f"年足データを元のExcelファイルに追加しました: {excel_file} (シート: {sheet_name})")
            except Exception as e:
                print(f"エラー: {e}")
        
        except Exception as e:
            print(f"処理中にエラーが発生しました: {e}")
        
        print("=" * 50)
    
    print("年足データのExcelファイルへの追加が完了しました。")

if __name__ == "__main__":
    main()
