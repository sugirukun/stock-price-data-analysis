import os
import pandas as pd
import glob
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def add_quarterly_to_excel(quarterly_file_path, excel_dir):
    """
    四半期足データをExcelファイルに追加する関数
    
    Parameters:
    quarterly_file_path (str): 四半期足データのCSVファイルパス
    excel_dir (str): Excelファイルが保存されているディレクトリパス
    
    Returns:
    bool: 処理が成功したかどうか
    """
    try:
        # CSVファイル名から銘柄名を取得
        file_name = os.path.basename(quarterly_file_path)
        ticker_name = file_name.replace('_四半期足.csv', '')
        
        # 四半期足データを読み込む
        df_quarterly = pd.read_csv(quarterly_file_path)
        
        # Dateカラムをdatetime型に変換
        df_quarterly['Date'] = pd.to_datetime(df_quarterly['Date'])
        
        # 対応するExcelファイルを探す
        excel_files = glob.glob(os.path.join(excel_dir, f"{ticker_name}*.xlsx"))
        
        if not excel_files:
            print(f"警告: {ticker_name}に対応するExcelファイルが見つかりませんでした。新規作成します。")
            # ティッカーシンボルまたは銘柄コードから銘柄名を抽出
            if "_" in ticker_name:
                # ファイル名に_が含まれる場合、最後の部分を銘柄名と仮定
                stock_name = ticker_name.split("_")[-1]
            else:
                # _がない場合はそのままを銘柄名として使用
                stock_name = ticker_name
                
            # シート名を作成（銘柄名_四半期足）
            sheet_name = f"{stock_name}_四半期足"
            
            # シート名の長さ制限（31文字まで）
            sheet_name = sheet_name[:31]
            
            # 新規Excelファイルを作成
            excel_path = os.path.join(excel_dir, f"{ticker_name}.xlsx")
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # 既存のシートに月足データを書き込む
                df_quarterly.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"新規Excelファイル作成: {excel_path}")
            return True
        
        # 最初に見つかったExcelファイルを使用
        excel_path = excel_files[0]
        print(f"Excelファイル検出: {excel_path}")
        
        # Excelファイルを開く
        book = load_workbook(excel_path)
        
        # ティッカーシンボルまたは銘柄コードから銘柄名を抽出
        # ファイル名のパターンは通常「コード_銘柄名_四半期足.csv」の形式と想定
        # または単に「銘柄名_四半期足.csv」の場合もある
        
        # ファイル名から銘柄名を抽出
        if "_" in ticker_name:
            # ファイル名に_が含まれる場合、最後の部分を銘柄名と仮定
            stock_name = ticker_name.split("_")[-1]
        else:
            # _がない場合はそのままを銘柄名として使用
            stock_name = ticker_name
            
        # シート名を作成（銘柄名_四半期足）
        sheet_name = f"{stock_name}_四半期足"
        
        # シート名の長さ制限（31文字まで）
        sheet_name = sheet_name[:31]
        
        # '銘柄名_四半期足'シートが存在するか確認し、なければ作成
        if sheet_name not in book.sheetnames:
            sheet = book.create_sheet(sheet_name)
        else:
            sheet = book[sheet_name]
            # シートをクリア
            for row in sheet.rows:
                for cell in row:
                    cell.value = None
        
        # DataFrameをシートに書き込む
        rows = dataframe_to_rows(df_quarterly, index=False, header=True)
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)
        
        # Excelファイルを保存
        book.save(excel_path)
        print(f"四半期足データをシート「{sheet_name}」に追加完了: {excel_path}")
        return True
        
    except Exception as e:
        print(f"エラー ({file_name}): {e}")
        return False

def main():
    """
    メイン関数：全ての四半期足CSVファイルをExcelに追加
    """
    # 株価データフォルダのパス
    data_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), '株価データ')
    # Excelフォルダのパス（株価データフォルダと同じと仮定）
    excel_folder = data_folder
    
    # 四半期足CSVファイルを検索
    quarterly_files = glob.glob(os.path.join(data_folder, '*_四半期足.csv'))
    
    print(f"処理対象ファイル数: {len(quarterly_files)}")
    
    success_count = 0
    # 各四半期足CSVファイルをExcelに追加
    for quarterly_file in quarterly_files:
        if add_quarterly_to_excel(quarterly_file, excel_folder):
            success_count += 1
    
    print(f"処理完了: {success_count}/{len(quarterly_files)} ファイルを処理しました")

if __name__ == "__main__":
    main()
